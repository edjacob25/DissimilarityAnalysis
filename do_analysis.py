import argparse
import configparser
import json
import multiprocessing
import subprocess
from dataclasses import astuple
from datetime import datetime
from shutil import copyfile, rmtree
from typing import Tuple
from zipfile import ZipFile, ZIP_BZIP2

import git
import math
import requests
import time
from itertools import product
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sty import fg, ef, rs, RgbFg

from common import format_time_lapse
from data_types import *


fg.set_style('orange', RgbFg(255, 150, 50))


def get_number_of_clusters(filepath: Path):
    with filepath.open("r") as file:
        for line in file:
            line_upper = line.upper()
            if line_upper.startswith("@ATTRIBUTE") and ("CLASS" in line_upper or "CLUSTER" in line_upper):
                clusters = line.split(" ", 2)[-1]
                return len(clusters.split(","))
            if line_upper.startswith("@DATA"):
                raise Exception("Could not found Class or Cluster attribute")


def remove_attribute(filepath: Path, attribute: str):
    attributes = []
    data = []
    permitted_indexes = []
    with filepath.open("r") as file:
        relation_name = file.readline()
        data_processing = False
        i = 0
        for line in file:
            if not data_processing:
                line_upper = line.upper()
                if line_upper.startswith("@ATTRIBUTE"):
                    if attribute.upper() != line_upper.split(" ")[1]:
                        attributes.append(line)
                        permitted_indexes.append(i)
                    i += 1
                    continue
                if line_upper.startswith("@DATA"):
                    data_processing = True
            else:
                line_data = [x.lstrip() for x in line.split(',')]
                data.append(line_data)

    with filepath.open("w") as new_file:
        new_file.write(relation_name)
        for item in attributes:
            new_file.write(item)
        new_file.write("\n")
        new_file.write("@DATA\n")
        for datapoint in data:
            points = [x for i, x in enumerate(datapoint) if i in permitted_indexes]
            result = ",".join(points)
            new_file.write(result)


# TODO: Add option to read classpath from the config file
def cluster_dataset(exp_parameters: ExperimentParameters, set_parameters: ExperimentSetParameters,
                    parameters: GeneralParameters, start_mode: str = "1") -> Experiment:
    _, measure, strategy, weight, learning_based, auc_type = astuple(exp_parameters)
    filepath = exp_parameters.filepath
    clustered_name = filepath.name.replace(".arff", f"_{measure}_{strategy}_{weight}_{auc_type}_clustered.arff")
    clustered_file_path = filepath.with_name(clustered_name)

    command = ["java", "-Xmx8192m"]

    if parameters.classpath is not None:
        java_classpath = parameters.classpath
    else:
        # Default classpath
        java_classpath = "/mnt/c/Program Files/Weka-3-9/weka.jar:/home/jacob/wekafiles/packages" \
                         "/SimmilarityMeasuresForCategoricalData/DissimilarityMeasures-0.1.jar"
    command.append("-cp")
    command.append(java_classpath)

    command.append("weka.filters.unsupervised.attribute.AddCluster")
    command.append("-W")

    num_classes = get_number_of_clusters(filepath)
    if parameters.verbose:
        print(f"Number of clusters for {filepath} is {num_classes}")
    num_procs = multiprocessing.cpu_count()
    if learning_based:
        distance_function = f"weka.core.{measure} -R first-last -S {strategy} -W {weight}"
        if parameters.normalize_dissimilarity:
            distance_function = f"{distance_function} -n"
    else:
        distance_function = f"weka.core.{measure}"

    if exp_parameters.auc_type is not None:
        distance_function = f"{distance_function} -a {auc_type}"

    if not set_parameters.initial:
        distance_function = f"\"{distance_function} -w {set_parameters.weight} -o {set_parameters.strategy} " \
                            f"-t {set_parameters.multiplier}\""
    else:
        distance_function = f"\"{distance_function}\""

    clusterer = f"weka.clusterers.CategoricalKMeans -init {start_mode} -max-candidates 100 -periodic-pruning 10000 " \
                f"-min-density 2.0 -t1 -1.25 -t2 -1.0 -N {num_classes} -M -A {distance_function} -I 500 " \
                f"-num-slots {math.floor(num_procs / 3)} -S 10"
    command.append(clusterer)
    command.append("-i")
    command.append(str(filepath))
    command.append("-o")
    command.append(str(clustered_file_path))
    command.append("-I")
    command.append("Last")

    start_dt = datetime.now()
    start = time.time()
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    end = time.time()

    identifier = f"measure {measure}"
    if exp_parameters.learning_based:
        identifier = f"{identifier}(strategy: {strategy}, weight: {weight})"
    if start_mode == "0":
        identifier = f"{identifier} classic mode"

    print(result.stderr.decode("utf-8"))
    if parameters.verbose:
        print(result.stdout.decode("utf-8"))
        print(f"Clustering dataset {filepath} with {identifier} took {end - start}")

    if "Exception" not in result.stderr.decode("utf-8"):
        remove_attribute(clustered_file_path, "Class")
        number_of_clusters = get_number_of_clusters(clustered_file_path)
        print(f"Finished clustering dataset {filepath} with {identifier}")
    else:
        if clustered_file_path.exists():
            clustered_file_path.unlink()

        if start_mode == "1":
            print(f"{fg.orange}There was an error running weka with k-means++ mode, trying with classic mode{fg.rs}")
            return cluster_dataset(exp_parameters, set_parameters, parameters, start_mode="0")
        else:
            raise Exception(f"{fg.red}There was a error running weka with the file {filepath.name} and the" +
                            f" following command{ef.italic} {' '.join(result.args)}{rs.italic}")

    if start_mode == "1":
        return Experiment(method=distance_function.replace("\"", ""), command_sent=" ".join(command),
                          time_taken=end - start, k_means_plusplus=True, file_name=filepath.name,
                          number_of_classes=num_classes, number_of_clusters=number_of_clusters,
                          start_time=start_dt, comments="")
    else:
        return Experiment(method=distance_function.replace("\"", ""), command_sent=" ".join(command),
                          time_taken=end - start, k_means_plusplus=False, file_name=filepath.name,
                          number_of_classes=num_classes, number_of_clusters=number_of_clusters,
                          start_time=start_dt, comments="")


def copy_files(exp_params: ExperimentParameters) -> Tuple[Path, Path]:
    _, measure, strategy, weight, learning_based, auc = astuple(exp_params)
    filepath = exp_params.filepath
    path: Path = filepath.parent
    file, filename = filepath.name, filepath.stem

    if learning_based:
        new_folder_path = path / f"{filename}_{measure}_{strategy}_{weight}_{auc}"
    else:
        new_folder_path = path / f"{filename}_{measure}"

    new_folder_path.mkdir()
    new_filepath = new_folder_path / file
    copyfile(str(filepath), new_filepath)

    old_clustered_filepath = path / f"{filename}_{measure}_{strategy}_{weight}_{auc}_clustered.arff"
    new_clustered_filepath = new_folder_path / f"{filename}.clus"
    copyfile(str(old_clustered_filepath), new_clustered_filepath)
    old_clustered_filepath.unlink()
    return new_filepath, new_clustered_filepath


def get_f_measure(filepath: Path, clustered_filepath: Path, exe_path: str = None, verbose: bool = False) -> str:
    command = ["MeasuresComparator.exe", "-c", str(clustered_filepath), "-r", str(filepath)]
    if exe_path is not None:
        command[0] = exe_path
    start = time.time()
    result = subprocess.run(command, stdout=subprocess.PIPE)
    end = time.time()
    text_result = result.stdout.decode('utf-8')

    if result.returncode != 0:
        print(f"{fg.red}Could not get F-Measure\nError ->{fg.rs} {text_result}")
        raise Exception("Could not calculate f-measure")
    else:
        if verbose:
            print(f"Calculating f-measure took {end - start}")
        print(f"{fg.green}Finished getting f-measure for {fg.blue}{filepath}{fg.green}, f-measure -> {fg.blue}"
              f"{text_result}{fg.rs}")
        return text_result


def send_notification(message: str, title: str):
    config = configparser.ConfigParser()
    config.read("config.ini")
    data = {"body": message, "title": title, "type": "note"}
    headers = {"Content-Type": "application/json", "Access-Token": config["SECRETS"]["Pushbullet_token"]}
    requests.post("https://api.pushbullet.com/v2/pushes", headers=headers, data=json.dumps(data))


def do_single_experiment(item: Path, strategy: str, weight: str, set_parameters: ExperimentSetParameters,
                         params: GeneralParameters, auc_type: str = None) -> Experiment:
    if weight is None:
        exp_params = ExperimentParameters(filepath=item, measure=strategy)
    else:
        exp_params = ExperimentParameters(filepath=item, measure="LearningBasedDissimilarity",
                                          strategy=strategy, weight_strategy=weight, learning_based=True,
                                          auc_type=auc_type)

    exp = cluster_dataset(exp_params, set_parameters, params)

    new_filepath, new_clustered_filepath = copy_files(exp_params)

    f_measure = get_f_measure(new_filepath, new_clustered_filepath,
                              exe_path=params.measure_calculator_path,
                              verbose=params.verbose)
    exp.f_score = f_measure
    return exp


def do_experiment_set(set_params: ExperimentSetParameters, params: GeneralParameters):
    if set_params.alternate:

        measures = ["Eskin", "Gambaryan", "Goodall", "Lin", "OccurenceFrequency", "InverseOccurenceFrequency",
                    "EuclideanDistance", "ManhattanDistance", "LinModified", "LinModified_Kappa",
                    "LinModified_MinusKappa", "LinModified_KappaMax"]
        if not set_params.initial:
            measures = ["EskinModified", "GambaryanModified", "GoodallModified", "OFModified", "IOFModified",
                        "LinModified"]

        measures = list(zip(measures, [None for _ in range(len(measures))]))
    else:
        strategies = ["A", "B", "C", "D", "E", "N"]
        weights = ["N", "K", "A"]
        measures = list(product(strategies, weights))

    engine = create_engine('sqlite:///Results/results.db')
    Base.metadata.create_all(engine)
    session_class = sessionmaker(bind=engine)
    session = session_class()
    code_directory = Path.cwd().parent
    repo = git.Repo(code_directory)

    description = f"{set_params.description}"
    if not set_params.initial:
        description = f"{description} using {set_params.weight} as decision weight, doing {set_params.strategy} when " \
                      f"weight is low and multiplying by {set_params.multiplier}"

    exp_set = ExperimentSet(time=datetime.now(), base_directory=str(params.directory), commit=repo.head.object.hexsha,
                            description=description)
    session.add(exp_set)
    session.commit()

    start = time.time()
    i = 0
    pool = multiprocessing.Pool(math.floor(multiprocessing.cpu_count() / 2))
    for item in params.directory.iterdir():
        if item.suffix == "arff" and "clustered" not in item.stem:
            try:
                sets = pool.starmap(do_single_experiment,
                                    [(item, strategy, weight, set_params, params) for strategy, weight in
                                     measures])
                exp_set.experiments.extend(sets)
                session.commit()
                i += 1
            except KeyboardInterrupt:
                session.rollback()
                print(f"{fg.orange}The analysis of the file {item} was requested to be finished by using Ctrl-C{fg.rs}")
                continue
            except Exception as exc:
                session.rollback()
                print(exc)
                print(f"{fg.red}Skipping file {item}{fg.rs}")
                continue
            finally:
                print("\n\n")

    end = time.time()

    exp_set.time_taken = end - start
    exp_set.number_of_datasets = i
    session.commit()
    # create_report(exp_set.id, base_path=root_dir)

    time_str = format_time_lapse(start, end)
    send_notification(f"It took {time_str} and processed {i} datasets", "Analysis finished")


def full_experiments(params: GeneralParameters):
    clean_experiments(params.directory)
    set_params = ExperimentSetParameters(initial=True, description="Initial Learning Based")
    do_experiment_set(set_params, params)
    set_params.alternate = True
    set_params.description = "Initial Other Measures"
    do_experiment_set(set_params, params)
    save_results(params.directory, f"Results_initial.zip")
    clean_experiments(params.directory)

    weights = ["K", "A"]
    strategies = ["B", "D", "M", "L"]
    multipliers = ["N", "I", "O"]
    for weight in weights:
        for strategy in strategies:
            for multiplier in multipliers:
                set_params = ExperimentSetParameters(strategy=strategy, multiplier=multiplier, weight=weight,
                                                     initial=False, description="Learning Based")
                do_experiment_set(set_params, params)
                set_params.alternate = True
                set_params.description = "Modified Measures"
                do_experiment_set(set_params, params)
                results_archive_name = f"Results_weight{weight}_strategy{strategy}_multiply_{multiplier}.zip"
                if params.normalize_dissimilarity:
                    results_archive_name = results_archive_name.replace(".zip", "_n.zip")
                save_results(params.directory, results_archive_name)
                clean_experiments(params.directory)


def clean_experiments(directory: Path):
    for item in directory.iterdir():
        if item.is_dir():
            rmtree(item.resolve())
            continue
        if "clustered" in item:
            item.unlink()


def save_results(base_directory: Path, filename: str):
    with ZipFile(base_directory / filename, 'w', compression=ZIP_BZIP2) as zipfile:
        for directory in base_directory.iterdir():
            if directory.is_dir():
                zipfile.write(directory.resolve(), arcname=directory.name)
                for file in directory.iterdir():
                    zipfile.write(file.resolve(), arcname=file.relative_to(base_directory))


def create_report(experiment_set: int, base_path: Path = Path(".")):
    wb = Workbook()
    ws = wb.active
    engine = create_engine('sqlite:///Results/results.db')
    session_class = sessionmaker(bind=engine)
    session = session_class()
    headers = []
    row = 1
    last = ""
    column = 2
    for experiment in session.query(Experiment).filter_by(set_id=experiment_set).order_by(Experiment.file_name):
        if last != experiment.file_name:
            column = 2
            row += 1
            last = experiment.file_name
            ws.cell(row=row, column=1, value=experiment.file_name.rsplit('/')[-1])
        if experiment.method not in headers:
            headers.append(experiment.method)
        ws.cell(row=row, column=column, value=experiment.f_score)
        column += 1

    for i, header in enumerate(headers):
        ws.cell(row=1, column=i + 2, value=header)
        ws.cell(row=1, column=i + column + 2, value=header)

    for i in range(2, row + 1):
        base = ord('A') + column - 2
        for j in range(column - 2):
            item = ord('B') + j
            ws.cell(row=i, column=j + column + 2, value=f"=RANK.AVG({chr(item)}{i},$B{i}:${chr(base)}{i})")

    start = chr(ord('B') + column)
    end = chr(ord('B') + column + column - 3)
    for i in range(column - 2):
        item = chr(ord('B') + i + column)
        ws.cell(row=row + 3, column=i + column + 2, value=f"=AVERAGE({item}2:{item}{row + 1})")
        ws.cell(row=row + 4, column=i + column + 2, value=f"=RANK({item}{row + 3},${start}{row + 3}:${end}{row + 3},1)")

    save_path = base_path / "results.xlsx"
    print(f"{fg.green}Saving report to {save_path}{fg.rs}")
    wb.save(str(save_path))


def main():
    parser = argparse.ArgumentParser(description='Does the analysis of a directory containing categorical datasets')
    parser.add_argument('directory', help="Directory in which the cleaned datasets are")
    parser.add_argument('-cp', help="Classpath for the weka invocation, needs to contain the weka.jar file and probably"
                                    " the jar of the measure ")
    parser.add_argument("-v", "--verbose", help="Show the output of the weka commands", action='store_true')
    parser.add_argument("-f", "--measure-calc", help="Path to the f-measure calculator", dest='measure_calculator_path')
    parser.add_argument("--alternate-analysis", help="Does the alternate analysis with the already known simmilarity "
                                                     "measures", action='store_true')
    parser.add_argument("-s", "--save", help="Path to the f-measure calculator", action='store_true')

    parser.add_argument("--full", help="Whether to do all combinations, overrides --alternate-analysis",
                        action='store_true')
    # TODO: Actually save the output of the commands

    args = parser.parse_args()

    # TODO: Read a single file
    directory = Path(args.directory)
    if not directory.is_dir():
        print(f"{fg.red}The selected path is not a directory{fg.rs}")
        exit(1)

    directory = directory.resolve()
    params = GeneralParameters(directory=directory, verbose=args.verbose, classpath=args.cp,
                               measure_calculator_path=args.measure_calculator_path)
    if args.full:
        print(f"{fg.red}DOING ALL EXPERIMENTS{fg.rs}")
        print(f"{fg.green}Go for a coffee{fg.rs}")
        full_experiments(params)
    else:
        results_archive_name = f"Results_single_alt_{args.alternate_analysis}.zip"
        save_results(params.directory, results_archive_name)
        clean_experiments(params.directory)
        set_params = ExperimentSetParameters(alternate=args.alternate_analysis)
        do_experiment_set(set_params, params)


if __name__ == "__main__":
    main()
