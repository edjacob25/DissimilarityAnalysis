# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.9.0
#   kernelspec:
#     display_name: thesis
#     language: python
#     name: thesis
# ---

# %% pycharm={"is_executing": false}
from pandas import DataFrame
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import random
from scipy.stats import wilcoxon

# %% pycharm={"is_executing": false}
loc = "/mnt/f/Datasets/CleanedDatasets2/Results_Other.xlsx"
wb = load_workbook(loc)
sheet = wb.active

# %% pycharm={"is_executing": false}
headers = [sheet.cell(1, x).value.split(" ", 1)[1] for x in range(2, 18, 2)]
headers.append("Learning Based")
headers

# %% pycharm={"is_executing": false}
data = [[sheet.cell(y, x).value for y in range(2, 31)] for x in range(3, 18, 2)]
data.append(
    [
        0.9566994905,
        0.4884400666,
        0.6525660753,
        0.926907599,
        0.9353886843,
        0.945910573,
        0.9069463015,
        0.6714044809,
        0.5670365095,
        0.6437994838,
        0.338203758,
        0.4083592594,
        0.4059928656,
        0.5114220381,
        0.2943754792,
        0.3979581892,
        0.4172253013,
        0.5566448569,
        0.4551629126,
        0.5562104583,
        0.5360223055,
        0.5547594428,
        0.4666725099,
        0.5604070425,
        0.444272697,
        0.2828712165,
        0.522356689,
        0.5395794511,
        0.688572228,
    ]
)
df = pd.DataFrame(data, index=headers)

df

# %% pycharm={"is_executing": false}
plt.isinteractive()

# %% pycharm={"is_executing": false}
means = df.mean(axis=1)
means

# %% pycharm={"is_executing": false}
all_colors = list(plt.cm.colors.cnames.keys())
random.seed(1000)
c = random.choices(all_colors, k=9)

# Plot Bars
plt.figure(figsize=(16, 10), dpi=80)
plt.bar(means.index, means, color=c, width=0.8)
for i, val in enumerate(means.values):
    plt.text(
        i,
        val,
        "{0:.5f}".format(val),
        horizontalalignment="center",
        verticalalignment="bottom",
        fontdict={"fontweight": 500, "size": 18},
    )

# Decoration
plt.gca().set_xticklabels(df.index, rotation=60, horizontalalignment="right", fontdict={"fontweight": 500, "size": 18})
for label in plt.gca().get_yticklabels():
    label.set_fontsize(18)  # Size here overrides font_prop

plt.title("Mean F1-Score in trained datasets", fontsize=22)
plt.ylabel("F1-Score", fontsize=20)
# plt.ylim(0, 0.7)
plt.savefig("bars.png", transparent=True, bbox_inches="tight")
plt.show()

# %% pycharm={"is_executing": false}
fig = plt.figure(figsize=(16, 10), dpi=80)
ax = fig.add_subplot(111)
# fig, ax = plt.subplots()
# box = df.boxplot(ax=ax)

bp = ax.boxplot(df, usermedians=means, autorange=True, widths=0.65, patch_artist=True)
ax.margins(y=0.05)

for label in plt.gca().get_yticklabels():
    label.set_fontsize(18)  # Size here overrides font_prop
for i, box in enumerate(bp["boxes"]):
    # change outline color
    # box.set( color='#7570b3', linewidth=2)
    # change fill color
    box.set_facecolor(c[i])
    pass
for median in bp["medians"]:
    median.set(color="black")

for i, val in enumerate(means.values):
    plt.text(
        i + 1,
        val,
        "{0:.4f}".format(val),
        horizontalalignment="center",
        verticalalignment="bottom",
        fontdict={"fontweight": 500, "size": 18},
    )

plt.ylabel("F1-Score", fontsize=20)
plt.gca().set_xticklabels(df.index, rotation=60, horizontalalignment="right", fontdict={"fontweight": 500, "size": 18})
plt.savefig("box.png", transparent=True, bbox_inches="tight")
plt.show()

# %% pycharm={"is_executing": false}
ls = df.index.values.tolist()

pairs = []
for i in df.index:
    idx = ls.index(i) + 1
    for j in ls[idx:]:
        pairs.append((i, j))
pairs = []
for i in range(0, 8):
    pairs.append(("Learning Based", ls[i]))

a = pd.DataFrame(index=df.index, columns=df.index)

for x, y in pairs:
    _, p = wilcoxon(df.loc[x], df.loc[y])
    if p < 0.05:
        best = x if means[x] > means[y] else y
        # print(f"{x} and {y} are statistically different and the better is {best}")
        # print(f"{p}")
        print(f"LB vs {y} -> {p}")
        a[y][x] = best
    else:
        print(f"{x} and {y} are not different")
        a[y][x] = "None"

# %% pycharm={"is_executing": false}
a
